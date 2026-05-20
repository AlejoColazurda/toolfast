# -*- coding: utf-8 -*-
import os
import sys
import subprocess
import re
import shutil
import urllib.request
import threading
from glob import glob

# =====================================================================
# CONFIGURACIÓN DE ACTUALIZACIÓN
# =====================================================================
VERSION = "1.0.0"
GITHUB_RAW_URL = "https://raw.githubusercontent.com/AlejoColazurda/toolfast/main/toolfast.py"
VERSION_REMOTA_DISPONIBLE = None
# =====================================================================

# Habilitar soporte de colores ANSI en la terminal de Windows
if os.name == 'nt':
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        kernel32.SetConsoleMode(kernel32.GetStdHandle(-11), 7)
    except Exception:
        pass

# Colores ANSI
C_GREEN = "\033[92m"
C_CYAN = "\033[96m"
C_YELLOW = "\033[93m"
C_RED = "\033[91m"
C_BOLD = "\033[1m"
C_END = "\033[0m"

# Intentar importar librerías necesarias o instalarlas dinámicamente
LIBS_REQUIRED = ["pypdf", "pillow", "openpyxl", "send2trash"]
LIBS_TO_INSTALL = []

for lib in LIBS_REQUIRED:
    try:
        if lib == "pillow":
            import PIL
        elif lib == "pypdf":
            import pypdf
        elif lib == "openpyxl":
            import openpyxl
        elif lib == "send2trash":
            import send2trash
    except ImportError:
        LIBS_TO_INSTALL.append(lib)

if LIBS_TO_INSTALL:
    print(f"{C_YELLOW}Instalando librerías requeridas de Python ({', '.join(LIBS_TO_INSTALL)})...{C_END}")
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install"] + LIBS_TO_INSTALL)
        print(f"{C_GREEN}Librerías instaladas correctamente.{C_END}\n")
    except Exception as e:
        print(f"{C_RED}Error al instalar librerías automáticamente: {e}{C_END}")
        print(f"Por favor, instálalas ejecutando: pip install {' '.join(LIBS_TO_INSTALL)}")

# Importaciones seguras de dependencias
import pypdf
from PIL import Image
import csv
try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import send2trash
except ImportError:
    send2trash = None

# =====================================================================
# SISTEMA DE AUTO-ACTUALIZACIÓN
# =====================================================================
def parse_version(v_str):
    try:
        return tuple(map(int, v_str.split('.')))
    except Exception:
        return (0, 0, 0)

def verificar_actualizacion_segundo_plano():
    global VERSION_REMOTA_DISPONIBLE
    try:
        req = urllib.request.Request(
            GITHUB_RAW_URL, 
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req, timeout=3) as response:
            codigo_remoto = response.read().decode('utf-8')
        match = re.search(r'VERSION\s*=\s*["\']([^"\']+)["\']', codigo_remoto)
        if match:
            version_remota = match.group(1)
            if parse_version(version_remota) > parse_version(VERSION):
                VERSION_REMOTA_DISPONIBLE = version_remota
    except Exception:
        pass

def comprobar_actualizaciones(silencioso=False):
    if not silencioso:
        print(f"\n{C_YELLOW}Buscando actualizaciones en GitHub...{C_END}")
    try:
        req = urllib.request.Request(
            GITHUB_RAW_URL, 
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req, timeout=5) as response:
            codigo_remoto = response.read().decode('utf-8')
            
        match = re.search(r'VERSION\s*=\s*["\']([^"\']+)["\']', codigo_remoto)
        if match:
            version_remota = match.group(1)
            if parse_version(version_remota) > parse_version(VERSION):
                print(f"\n{C_GREEN}{C_BOLD}¡Nueva versión disponible! ({version_remota}){C_END}")
                print(f"Tu versión actual es: {VERSION}")
                confirmacion = input("¿Deseas actualizar el programa ahora? (S/N, por defecto S): ").strip().upper()
                if confirmacion in ('S', 'SI', ''):
                    ruta_local = os.path.abspath(__file__)
                    with open(ruta_local, 'w', encoding='utf-8') as f:
                        f.write(codigo_remoto)
                    print(f"\n{C_GREEN}{C_BOLD}¡Actualización exitosa!{C_END} Reinicia la herramienta.")
                    sys.exit(0)
            else:
                if not silencioso:
                    print(f"{C_GREEN}¡Ya tienes la última versión instalada! (v{VERSION}){C_END}")
    except Exception as e:
        if not silencioso:
            print(f"{C_RED}No se pudo comprobar actualizaciones: {e}{C_END}")

# =====================================================================
# AUXILIARES
# =====================================================================
def print_header(title=""):
    os.system('cls' if os.name == 'nt' else 'clear')
    print(f"{C_CYAN}{C_BOLD}" + "="*60)
    print(f"               TOOLFAST - NAVAJA SUIZA CLI v{VERSION}         ")
    print(f"                  Desarrollado por Alejo Colazurda           ")
    print("="*60 + f"{C_END}")
    
    global VERSION_REMOTA_DISPONIBLE
    if VERSION_REMOTA_DISPONIBLE:
        print(f"{C_YELLOW}{C_BOLD}[!] ¡NUEVA ACTUALIZACIÓN DISPONIBLE! (v{VERSION_REMOTA_DISPONIBLE}){C_END}")
        print(f"{C_YELLOW}    Selecciona la opción 5 en el menú principal para actualizar ahora.{C_END}")
        print(f"{C_YELLOW}" + "-"*60 + f"{C_END}")
        
    if title:
        print(f"{C_BOLD}--- {title} ---{C_END}\n")

def presionar_enter():
    input(f"\nPresiona {C_BOLD}Enter{C_END} para volver...")

def seleccionar_archivo_interactivo(extensiones, mensaje):
    """
    Permite al usuario ingresar una ruta completa de archivo o palabras clave.
    Busca coincidencias recursivamente (max 2 niveles) en la carpeta actual y Descargas,
    mostrando resultados numerados para fácil selección.
    """
    while True:
        entrada = input(mensaje).strip().strip('"')
        
        # Si es un archivo válido directo
        if entrada and os.path.isfile(entrada):
            ext = os.path.splitext(entrada)[1].lower()
            if ext in extensiones:
                return os.path.abspath(entrada)
            else:
                print(f"{C_RED}El archivo seleccionado no tiene una extensión válida ({', '.join(extensiones)}).{C_END}")
                continue
                
        # Buscar coincidencias en la carpeta actual y en Descargas
        carpetas_busqueda = [os.getcwd()]
        user_profile = os.environ.get("USERPROFILE")
        if user_profile:
            downloads_dir = os.path.join(user_profile, "Downloads")
            if os.path.exists(downloads_dir) and downloads_dir not in carpetas_busqueda:
                carpetas_busqueda.append(downloads_dir)
                
        coincidencias = []
        palabras = entrada.lower().split()
        
        for carpeta in carpetas_busqueda:
            try:
                for root, dirs, files in os.walk(carpeta):
                    # No ir más profundo de 2 niveles para que la búsqueda sea rápida
                    depth = root[len(carpeta):].count(os.sep)
                    if depth > 2:
                        continue
                        
                    for f in files:
                        ext_f = os.path.splitext(f)[1].lower()
                        if ext_f in extensiones:
                            # Comprobar si contiene las palabras clave
                            if all(p in f.lower() for p in palabras):
                                coincidencia_path = os.path.join(root, f)
                                if coincidencia_path not in coincidencias:
                                    coincidencias.append(coincidencia_path)
            except Exception:
                pass
                
        # Limitar resultados
        coincidencias = coincidencias[:15]
        
        if not coincidencias:
            print(f"{C_RED}No se encontró ningún archivo con esas palabras clave o extensión en esta carpeta ni en Descargas.{C_END}")
            print(f"Extensiones aceptadas: {', '.join(extensiones)}")
            reintentar = input("¿Deseas buscar de nuevo? (S/N, por defecto S): ").strip().upper()
            if reintentar in ("N", "NO"):
                return None
            continue
            
        print(f"\n{C_CYAN}--- Archivos encontrados ({len(coincidencias)}) ---{C_END}")
        for idx, coinc in enumerate(coincidencias, 1):
            base = os.path.basename(coinc)
            dir_name = os.path.basename(os.path.dirname(coinc))
            print(f"  {C_GREEN}{idx}.{C_END} ...\\{dir_name}\\{base}")
            
        print(f"  {C_RED}0.{C_END} Cancelar selección")
        
        seleccion = input(f"\nSelecciona el número de archivo (1-{len(coincidencias)}): ").strip()
        if seleccion == "0":
            return None
        if seleccion.isdigit() and 1 <= int(seleccion) <= len(coincidencias):
            return coincidencias[int(seleccion) - 1]
        else:
            print(f"{C_RED}Selección no válida.{C_END}")
            import time
            time.sleep(1.5)

# =====================================================================
# SECCIÓN 1: GESTIÓN DE PDFS
# =====================================================================
def menu_pdf():
    while True:
        print_header("GESTIÓN DE ARCHIVOS PDF")
        print(f" {C_GREEN}1.{C_END} Unificar carpeta completa de PDFs (orden alfabético)")
        print(f" {C_GREEN}2.{C_END} Separar o extraer páginas de un PDF")
        print(f" {C_GREEN}3.{C_END} Comprimir un PDF (reducir peso de streams)")
        print(f" {C_GREEN}4.{C_END} Extractor de Texto / OCR básico")
        print(f" {C_GREEN}5.{C_END} Volver al menú principal")
        
        opcion = input(f"\nSelecciona una opción (1-5): ").strip()
        if opcion == "1":
            pdf_unificar()
        elif opcion == "2":
            pdf_separar()
        elif opcion == "3":
            pdf_comprimir()
        elif opcion == "4":
            pdf_extraer_texto()
        elif opcion == "5":
            break
        else:
            print(f"{C_RED}Opción no válida.{C_END}")
            presionar_enter()

def pdf_unificar():
    print_header("UNIFICAR PDFs")
    carpeta = input("Ruta de la carpeta con los PDFs: ").strip().strip('"')
    if not os.path.exists(carpeta):
        print(f"{C_RED}La carpeta no existe.{C_END}")
        presionar_enter()
        return
    
    salida = input("Nombre o ruta del archivo resultante (ej: unificado.pdf): ").strip().strip('"')
    if not salida: salida = "unificado.pdf"
    if not salida.lower().endswith(".pdf"): salida += ".pdf"
    
    salida_abs = os.path.abspath(salida) if os.path.isabs(salida) else os.path.abspath(os.path.join(carpeta, salida))
    
    pdfs = glob(os.path.join(carpeta, "*.pdf"))
    pdfs = [p for p in pdfs if os.path.abspath(p) != salida_abs]
    
    if not pdfs:
        print(f"{C_YELLOW}No se encontraron PDFs en la carpeta.{C_END}")
        presionar_enter()
        return
        
    pdfs.sort(key=lambda x: os.path.basename(x).lower())
    
    print(f"\nArchivos encontrados ({len(pdfs)}):")
    for f in pdfs:
        print(f" - {os.path.basename(f)}")
        
    confirmar = input("\n¿Proceder con la unificación? (S/N, por defecto S): ").strip().upper()
    if confirmar not in ("S", "SI", ""):
        print(f"{C_YELLOW}Operación cancelada.{C_END}")
        presionar_enter()
        return
        
    # Pre-comprobación de encriptación
    encriptados = []
    for pdf in pdfs:
        try:
            reader = pypdf.PdfReader(pdf)
            if reader.is_encrypted:
                encriptados.append(os.path.basename(pdf))
        except Exception:
            pass
    if encriptados:
        print(f"\n{C_RED}Error: Se encontraron PDFs protegidos o encriptados:{C_END}")
        for enc in encriptados:
            print(f" - {enc}")
        print("Desbloquéalos antes de intentar unificarlos.")
        presionar_enter()
        return
        
    print(f"{C_YELLOW}Unificando...{C_END}")
    try:
        merger = pypdf.PdfMerger()
        for pdf in pdfs:
            merger.append(pdf)
        merger.write(salida_abs)
        merger.close()
        print(f"{C_GREEN}¡Éxito! PDF guardado en: {salida_abs}{C_END}")
    except Exception as e:
        print(f"{C_RED}Error al unificar: {e}{C_END}")
    presionar_enter()

def pdf_separar():
    print_header("SEPARAR O EXTRAER PÁGINAS")
    archivo = seleccionar_archivo_interactivo([".pdf"], "Escribe parte del nombre del PDF a separar (o la ruta completa): ")
    if not archivo:
        return
        
    print("\n¿Qué deseas hacer?")
    print(" 1. Guardar cada página como un archivo PDF individual")
    print(" 2. Extraer un rango de páginas específico (ej: 1-3, o 2,4,6)")
    subop = input("Selecciona (1 o 2): ").strip()
    
    try:
        reader = pypdf.PdfReader(archivo)
        if reader.is_encrypted:
            print(f"{C_RED}Error: El PDF está protegido con contraseña. Desbloquéalo antes de procesarlo.{C_END}")
            presionar_enter()
            return
            
        total_paginas = len(reader.pages)
        print(f"Páginas detectadas en el documento: {total_paginas}")
        
        if subop == "1":
            prefijo = input("Prefijo para los archivos resultantes (por defecto 'pagina'): ").strip()
            if not prefijo: prefijo = "pagina"
            
            carpeta_salida = os.path.join(os.path.dirname(archivo), "paginas_extraidas")
            os.makedirs(carpeta_salida, exist_ok=True)
            
            for idx in range(total_paginas):
                writer = pypdf.PdfWriter()
                writer.add_page(reader.pages[idx])
                salida = os.path.join(carpeta_salida, f"{prefijo}_{idx+1}.pdf")
                with open(salida, "wb") as f:
                    writer.write(f)
            print(f"{C_GREEN}¡Completado! Las páginas fueron guardadas en: {carpeta_salida}{C_END}")
            
        elif subop == "2":
            rango = input("Ingresa las páginas o rangos (ej: '1-3, 5' para páginas 1, 2, 3 y 5): ").strip()
            paginas_a_extraer = set()
            try:
                partes = rango.split(",")
                for parte in partes:
                    parte = parte.strip()
                    if not parte:
                        continue
                    if "-" in parte:
                        subpartes = parte.split("-")
                        if len(subpartes) != 2:
                            raise ValueError()
                        inicio, fin = map(int, subpartes)
                        for p in range(inicio, fin + 1):
                            paginas_a_extraer.add(p - 1)
                    else:
                        paginas_a_extraer.add(int(parte) - 1)
            except ValueError:
                print(f"{C_RED}Error: El formato de rango no es válido. Usa números separados por comas o guiones (ej: 1-3, 5).{C_END}")
                presionar_enter()
                return
            
            # Validar e insertar
            writer = pypdf.PdfWriter()
            agregadas = 0
            for idx in sorted(list(paginas_a_extraer)):
                if 0 <= idx < total_paginas:
                    writer.add_page(reader.pages[idx])
                    agregadas += 1
            
            if agregadas > 0:
                ruta_salida = os.path.join(os.path.dirname(archivo), "extraido_rango.pdf")
                with open(ruta_salida, "wb") as f:
                    writer.write(f)
                print(f"{C_GREEN}¡Éxito! Rango guardado en: {ruta_salida}{C_END}")
            else:
                print(f"{C_RED}No se seleccionaron páginas válidas.{C_END}")
        else:
            print(f"{C_RED}Opción incorrecta.{C_END}")
    except Exception as e:
        print(f"{C_RED}Error: {e}{C_END}")
    presionar_enter()

def pdf_comprimir():
    print_header("COMPRIMIR PDF")
    archivo = seleccionar_archivo_interactivo([".pdf"], "Escribe parte del nombre del PDF a comprimir (o la ruta completa): ")
    if not archivo:
        return
        
    salida = os.path.join(os.path.dirname(archivo), "comprimido_" + os.path.basename(archivo))
    print(f"{C_YELLOW}Comprimiendo flujos de contenido y removiendo metadatos redundantes...{C_END}")
    try:
        reader = pypdf.PdfReader(archivo)
        if reader.is_encrypted:
            print(f"{C_RED}Error: El PDF está protegido con contraseña. Desbloquéalo antes de comprimirlo.{C_END}")
            presionar_enter()
            return
        writer = pypdf.PdfWriter()
        
        for page in reader.pages:
            page.compress_content_streams() # Comprime flujos de texto/vectores
            writer.add_page(page)
            
        with open(salida, "wb") as f:
            writer.write(f)
            
        t_orig = os.path.getsize(archivo) / 1024
        t_comp = os.path.getsize(salida) / 1024
        ahorro = ((t_orig - t_comp) / t_orig) * 100 if t_orig > 0 else 0
        
        print(f"\n{C_GREEN}¡Completado!{C_END}")
        print(f"Tamaño original: {t_orig:.2f} KB")
        print(f"Tamaño comprimido: {t_comp:.2f} KB")
        print(f"Reducción de tamaño: {ahorro:.1f}%")
        print(f"Guardado en: {salida}")
    except Exception as e:
        print(f"{C_RED}Error al comprimir: {e}{C_END}")
    presionar_enter()

def pdf_extraer_texto():
    print_header("EXTRACTOR DE TEXTO / OCR")
    archivo = seleccionar_archivo_interactivo([".pdf"], "Escribe parte del nombre del PDF a extraer (o la ruta completa): ")
    if not archivo:
        return
        
    salida = os.path.join(os.path.dirname(archivo), os.path.splitext(os.path.basename(archivo))[0] + "_texto.txt")
    print(f"{C_YELLOW}Extrayendo texto del PDF...{C_END}")
    try:
        reader = pypdf.PdfReader(archivo)
        if reader.is_encrypted:
            print(f"{C_RED}Error: El PDF está protegido con contraseña. Desbloquéalo antes de extraer su texto.{C_END}")
            presionar_enter()
            return
        texto = []
        for idx, page in enumerate(reader.pages, 1):
            t_page = page.extract_text()
            if t_page:
                texto.append(f"--- PAGINA {idx} ---\n{t_page}\n")
                
        if not texto:
            print(f"{C_YELLOW}Aviso: No se pudo extraer texto embebido. Es posible que el PDF esté escaneado como imagen.{C_END}")
            # Intentar ver si se puede sugerir OCR o simplemente guardar vacío
            texto.append("[PDF escaneado o sin caracteres de texto embebido]")
            
        with open(salida, "w", encoding="utf-8") as f:
            f.writelines(texto)
        print(f"{C_GREEN}¡Texto extraído! Guardado en: {salida}{C_END}")
    except Exception as e:
        print(f"{C_RED}Error al extraer texto: {e}{C_END}")
    presionar_enter()

# =====================================================================
# SECCIÓN 2: CONVERSOR DE FORMATOS
# =====================================================================
def conversor_formatos():
    print_header("CONVERSOR DE FORMATOS DE ARCHIVO")
    extensiones_validas = [".png", ".jpg", ".jpeg", ".webp", ".bmp", ".csv", ".xlsx"]
    origen = seleccionar_archivo_interactivo(extensiones_validas, "Escribe parte del nombre del archivo a convertir (o la ruta completa): ")
    if not origen:
        return
        
    ext_orig = os.path.splitext(origen)[1].lower().replace(".", "")
    print(f"Archivo de origen detectado: Tipo '{ext_orig}'")
    
    destino = input("Formato o extensión final al que deseas convertir (ej: png, pdf, csv, xlsx): ").strip().lower().replace(".", "")
    if not destino:
        print(f"{C_RED}Formato final no válido.{C_END}")
        presionar_enter()
        return
        
    # Verificar compatibilidad y realizar conversión
    realizar_conversion(origen, ext_orig, destino)

def realizar_conversion(ruta_orig, ext_orig, ext_dest):
    # Definición de tipos de archivos
    IMAGENES = ["png", "jpg", "jpeg", "webp", "bmp"]
    
    # 1. Imagen a Imagen
    if ext_orig in IMAGENES and ext_dest in IMAGENES:
        try:
            print(f"{C_YELLOW}Convirtiendo imagen...{C_END}")
            img = Image.open(ruta_orig)
            salida = os.path.splitext(ruta_orig)[0] + "." + ext_dest
            
            # Manejar el canal alfa (transparencia PNG) para JPG
            if ext_dest in ["jpg", "jpeg"] and img.mode in ("RGBA", "P"):
                img = img.convert("RGB")
            img.save(salida)
            print(f"{C_GREEN}¡Éxito! Imagen convertida en: {salida}{C_END}")
        except PermissionError:
            print(f"{C_RED}Error de acceso: El archivo o la carpeta de destino están bloqueados o no tienes permisos de escritura.{C_END}")
        except Exception as e:
            print(f"{C_RED}Error al convertir imagen: {e}{C_END}")
        presionar_enter()
        return

    # 2. Imagen a PDF
    if ext_orig in IMAGENES and ext_dest == "pdf":
        try:
            print(f"{C_YELLOW}Convirtiendo imagen a documento PDF...{C_END}")
            img = Image.open(ruta_orig)
            salida = os.path.splitext(ruta_orig)[0] + ".pdf"
            if img.mode in ("RGBA", "P"):
                img = img.convert("RGB")
            img.save(salida, "PDF", resolution=100.0)
            print(f"{C_GREEN}¡Éxito! PDF de imagen creado en: {salida}{C_END}")
        except PermissionError:
            print(f"{C_RED}Error de acceso: El archivo o la carpeta de destino están bloqueados o no tienes permisos de escritura.{C_END}")
        except Exception as e:
            print(f"{C_RED}Error al convertir imagen a PDF: {e}{C_END}")
        presionar_enter()
        return

    # 3. CSV a Excel
    if ext_orig == "csv" and ext_dest == "xlsx":
        if not openpyxl:
            print(f"{C_RED}Se requiere la librería 'openpyxl' para crear archivos Excel.{C_END}")
            presionar_enter()
            return
        try:
            print(f"{C_YELLOW}Convirtiendo CSV a planilla de Excel...{C_END}")
            salida = os.path.splitext(ruta_orig)[0] + ".xlsx"
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Detectar delimitador leyendo la primera línea
            with open(ruta_orig, "r", encoding="utf-8-sig") as f:
                header = f.readline()
                delimitador = ";" if ";" in header else ","
            
            with open(ruta_orig, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f, delimiter=delimitador)
                for row in reader:
                    ws.append(row)
                    
            wb.save(salida)
            print(f"{C_GREEN}¡Excel generado con éxito en: {salida}{C_END}")
        except PermissionError:
            print(f"{C_RED}Error de acceso: El archivo de destino está abierto en otro programa (como Excel) o está bloqueado.{C_END}")
            print(f"{C_YELLOW}Por favor, cierra Excel o el archivo bloqueado e inténtalo de nuevo.{C_END}")
        except Exception as e:
            print(f"{C_RED}Error al convertir CSV a Excel: {e}{C_END}")
        presionar_enter()
        return

    # 4. Excel a CSV
    if ext_orig == "xlsx" and ext_dest == "csv":
        if not openpyxl:
            print(f"{C_RED}Se requiere la librería 'openpyxl' para abrir archivos Excel.{C_END}")
            presionar_enter()
            return
        try:
            print(f"{C_YELLOW}Convirtiendo Excel a archivo CSV...{C_END}")
            salida = os.path.splitext(ruta_orig)[0] + ".csv"
            
            wb = openpyxl.load_workbook(ruta_orig, data_only=True)
            sh = wb.active
            
            with open(salida, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f, delimiter=";")
                for r in sh.iter_rows(values_only=True):
                    # Filtrar filas vacías
                    if any(x is not None for x in r):
                        writer.writerow(r)
            print(f"{C_GREEN}¡CSV generado con éxito en: {salida}{C_END}")
        except PermissionError:
            print(f"{C_RED}Error de acceso: El archivo de destino está abierto en otro programa (como Excel) o está bloqueado.{C_END}")
            print(f"{C_YELLOW}Por favor, cierra Excel o el archivo bloqueado e inténtalo de nuevo.{C_END}")
        except Exception as e:
            print(f"{C_RED}Error al convertir Excel a CSV: {e}{C_END}")
        presionar_enter()
        return

    # 5. CONVERSIÓN IMPOSIBLE / NO SOPORTADA
    print(f"\n{C_RED}{C_BOLD}¡Ese tipo de conversión no es posible!{C_END}")
    
    # Sistema de sugerencias inteligentes
    if ext_orig == "pdf" and ext_dest in ["xlsx", "csv"]:
        print(f"No se puede convertir PDF a Excel de forma directa, pero puedes probar:\n"
              f" - {C_CYAN}Extraer el Texto (.txt){C_END} del PDF usando la opción 1 (menú PDF > Opción 4).\n"
              f" - Si el PDF tiene tablas, usar herramientas dedicadas online como IlovePDF.")
    elif ext_orig in IMAGENES and ext_dest in ["mp3", "mp4", "wav"]:
        print(f"No se puede convertir una imagen fija en archivo de audio/video, pero puedes:\n"
              f" - Convertirla a otro formato de imagen ({', '.join(IMAGENES)})\n"
              f" - Empaquetarla en un {C_CYAN}Documento PDF{C_END} (.pdf).")
    elif ext_orig in ["xlsx", "csv"] and ext_dest in IMAGENES:
        print(f"No se puede transformar una planilla estructurada en imagen, pero puedes:\n"
              f" - Guardarla como {C_CYAN}PDF (.pdf){C_END} desde Excel para visualizarla.")
    else:
        print(f"Formatos '{ext_orig}' y '{ext_dest}' no son compatibles en esta herramienta.\n"
              f"Prueba con combinaciones soportadas de {C_GREEN}PDF, Hojas de cálculo (Excel/CSV) e Imágenes (PNG/JPG){C_END}.")
              
    presionar_enter()

# =====================================================================
# SECCIÓN 3: GESTIÓN Y ORGANIZACIÓN
# =====================================================================
def menu_organizacion():
    while True:
        print_header("ORGANIZACIÓN Y HERRAMIENTAS DE OFICINA")
        print(f" {C_GREEN}1.{C_END} Renombrado masivo de archivos")
        print(f" {C_GREEN}2.{C_END} Organizador automático de carpetas (Limpiador)")
        print(f" {C_GREEN}3.{C_END} ¿Computadora lenta? Limpiar archivos basura de Windows")
        print(f" {C_GREEN}4.{C_END} Volver al menú principal")
        
        opcion = input(f"\nSelecciona una opción (1-4): ").strip()
        if opcion == "1":
            archivos_renombrar_masivo()
        elif opcion == "2":
            archivos_organizar_carpeta()
        elif opcion == "3":
            limpiar_archivos_basura()
        elif opcion == "4":
            break
        else:
            print(f"{C_RED}Opción no válida.{C_END}")
            presionar_enter()

def limpiar_archivos_basura():
    import ctypes
    print_header("LIMPIADOR DE ARCHIVOS BASURA (TEMPORALES)")
    print(f"{C_YELLOW}Este módulo liberará espacio en disco eliminando archivos temporales e inservibles.{C_END}")
    print("Se limpiarán las siguientes ubicaciones:")
    print(" 1. Carpeta de archivos temporales del usuario (%TEMP%)")
    print(" 2. Carpeta de archivos temporales del sistema (C:\\Windows\\Temp)")
    print(" 3. Vaciado de la Papelera de Reciclaje (opcional)")
    print(f"\n{C_BOLD}Nota:{C_END} Los archivos que estén siendo usados por programas abiertos se omitirán automáticamente.")
    
    confirmar = input("\n¿Proceder con la limpieza? (S/N, por defecto S): ").strip().upper()
    if confirmar not in ("S", "SI", ""):
        print(f"{C_YELLOW}Limpieza cancelada.{C_END}")
        presionar_enter()
        return
        
    vaciar_papelera = input("¿Deseas vaciar también la Papelera de Reciclaje? (S/N, por defecto S): ").strip().upper()
    vaciar_papelera = vaciar_papelera in ("S", "SI", "")
    
    print(f"\n{C_YELLOW}Iniciando limpieza...{C_END}\n")
    
    # 1. Limpiar %TEMP% del usuario
    temp_usuario = os.environ.get('TEMP')
    eliminados = 0
    liberado_bytes = 0
    errores = 0
    
    if temp_usuario and os.path.exists(temp_usuario):
        print(f"Limpiando temporales del usuario: {temp_usuario}...")
        for root, dirs, files in os.walk(temp_usuario):
            for f in files:
                fp = os.path.join(root, f)
                try:
                    sz = os.path.getsize(fp)
                    os.remove(fp)
                    eliminados += 1
                    liberado_bytes += sz
                except Exception:
                    errores += 1
            for d in dirs:
                dp = os.path.join(root, d)
                try:
                    shutil.rmtree(dp)
                except Exception:
                    pass
                    
    # 2. Limpiar C:\Windows\Temp
    temp_sistema = "C:\\Windows\\Temp"
    if os.path.exists(temp_sistema):
        print(f"Limpiando temporales del sistema: {temp_sistema}...")
        for root, dirs, files in os.walk(temp_sistema):
            for f in files:
                fp = os.path.join(root, f)
                try:
                    sz = os.path.getsize(fp)
                    os.remove(fp)
                    eliminados += 1
                    liberado_bytes += sz
                except Exception:
                    errores += 1
            for d in dirs:
                dp = os.path.join(root, d)
                try:
                    shutil.rmtree(dp)
                except Exception:
                    pass
                    
    # 3. Vaciar la papelera de reciclaje usando ctypes shell32
    if vaciar_papelera and os.name == 'nt':
        print("Vaciando Papelera de Reciclaje...")
        try:
            res = ctypes.windll.shell32.SHEmptyRecycleBinW(None, None, 7)
            if res == 0:
                print(f"{C_GREEN}Papelera vaciada con éxito.{C_END}")
            else:
                print(f"{C_YELLOW}Nota: La papelera ya estaba vacía o requiere permisos adicionales.{C_END}")
        except Exception as e:
            print(f"No se pudo vaciar la papelera: {e}")
            
    liberado_mb = liberado_bytes / (1024 * 1024)
    print(f"\n{C_GREEN}{C_BOLD}¡LIMPIEZA COMPLETADA CON ÉXITO!{C_END}")
    print(f" Archivos eliminados: {C_GREEN}{eliminados}{C_END}")
    print(f" Espacio liberado: {C_GREEN}{liberado_mb:.2f} MB{C_END}")
    if errores > 0:
        print(f" Archivos omitidos (en uso activo por el sistema): {C_YELLOW}{errores}{C_END}")
    presionar_enter()

def archivos_renombrar_masivo():
    print_header("RENOMBRADO MASIVO DE ARCHIVOS")
    carpeta = input("Ruta de la carpeta con los archivos: ").strip().strip('"')
    if not os.path.exists(carpeta):
        print(f"{C_RED}La carpeta no existe.{C_END}")
        presionar_enter()
        return
        
    filtro = input("Extensión de archivos a renombrar (ej: *.pdf, *.png, o presiona Enter para todos): ").strip()
    if not filtro: filtro = "*.*"
    if not filtro.startswith("*"): filtro = "*" + filtro
    
    archivos = glob(os.path.join(carpeta, filtro))
    archivos = [a for a in archivos if os.path.isfile(a)]
    
    if not archivos:
        print(f"{C_YELLOW}No se encontraron archivos coincidentes.{C_END}")
        presionar_enter()
        return
        
    print(f"\nSe encontraron {len(archivos)} archivos.")
    print("Elige la regla de renombrado:")
    print(" 1. Añadir un prefijo al nombre (ej: 'Factura_' -> 'Factura_archivo.pdf')")
    print(" 2. Reemplazar texto en el nombre (ej: reemplazar '2023' por '2024')")
    print(" 3. Numeración secuencial (ej: 'Foto_1.jpg', 'Foto_2.jpg'...)")
    regla = input("Selecciona (1-3): ").strip()
    
    if regla == "1":
        prefijo = input("Ingresa el prefijo: ")
        for a in archivos:
            base = os.path.basename(a)
            nuevo = os.path.join(carpeta, prefijo + base)
            os.rename(a, nuevo)
        print(f"{C_GREEN}¡Archivos renombrados con éxito!{C_END}")
        
    elif regla == "2":
        buscar = input("Texto a buscar: ")
        reemplazar = input("Texto de reemplazo: ")
        renombrados = 0
        for a in archivos:
            base = os.path.basename(a)
            if buscar in base:
                nuevo = os.path.join(carpeta, base.replace(buscar, reemplazar))
                os.rename(a, nuevo)
                renombrados += 1
        print(f"{C_GREEN}¡Completado! Se renombraron {renombrados} archivos.{C_END}")
        
    elif regla == "3":
        nombre_base = input("Nombre base para la serie (ej: Remito): ").strip()
        if not nombre_base: nombre_base = "Archivo"
        ext = os.path.splitext(archivos[0])[1]
        
        # Ordenar archivos antes de numerar
        archivos.sort(key=lambda x: os.path.basename(x).lower())
        
        for idx, a in enumerate(archivos, 1):
            ext_actual = os.path.splitext(a)[1]
            nuevo = os.path.join(carpeta, f"{nombre_base}_{idx}{ext_actual}")
            os.rename(a, nuevo)
        print(f"{C_GREEN}¡Archivos numerados correctamente!{C_END}")
    else:
        print(f"{C_RED}Opción no válida.{C_END}")
    presionar_enter()

def archivos_organizar_carpeta():
    print_header("ORGANIZADOR AUTOMÁTICO DE CARPETAS")
    print("Navega por tus directorios para seleccionar cuál deseas organizar.")
    
    ruta = input("\nRuta inicial (deja vacío para la carpeta actual): ").strip().strip('"')
    if not ruta:
        ruta = os.getcwd()
        
    while True:
        ruta = os.path.abspath(ruta)
        if not os.path.exists(ruta) or not os.path.isdir(ruta):
            print(f"{C_RED}La ruta '{ruta}' no es una carpeta válida.{C_END}")
            presionar_enter()
            return
            
        print_header(f"Organizador - Carpeta Actual: {ruta}")
        
        try:
            items = os.listdir(ruta)
        except Exception as e:
            print(f"{C_RED}Error al acceder a la carpeta: {e}{C_END}")
            print(f"{C_YELLOW}Regresando al directorio padre...{C_END}")
            import time
            time.sleep(2)
            ruta = os.path.dirname(ruta)
            continue
            
        directorios = []
        archivos = []
        for item in items:
            if item.startswith('.') or item.startswith('$'):
                continue
            full_path = os.path.join(ruta, item)
            try:
                if os.path.isdir(full_path):
                    directorios.append(item)
                else:
                    archivos.append(item)
            except Exception:
                pass
                
        directorios.sort(key=lambda x: x.lower())
        archivos.sort(key=lambda x: x.lower())
        
        print(f"{C_BOLD}Opciones de acción:{C_END}")
        print(f"  {C_GREEN}[ O ]{C_END} ---> {C_BOLD}ORGANIZAR ARCHIVOS DE ESTA CARPETA{C_END}")
        print(f"  {C_YELLOW}[ .. ]{C_END} ---> Subir de nivel")
        print(f"  {C_RED}[ salir ]{C_END} -> Cancelar y salir")
        
        idx = 1
        mapeo_carpetas = {}
        
        if directorios:
            print(f"\n{C_CYAN}--- Subcarpetas disponibles ({len(directorios)}) ---{C_END}")
            print("Ingresa el número para ingresar a la subcarpeta:")
            for d in directorios:
                print(f"  {C_GREEN}{idx}.{C_END} [CARPETA] {d}")
                mapeo_carpetas[idx] = d
                idx += 1
                
        if archivos:
            print(f"\n{C_CYAN}--- Archivos sueltos a clasificar ({len(archivos)}) ---{C_END}")
            for file in archivos:
                try:
                    size = os.path.getsize(os.path.join(ruta, file)) / 1024
                    size_str = f"({size:.1f} KB)"
                except Exception:
                    size_str = ""
                print(f"   - {file} {C_YELLOW}{size_str}{C_END}")
                
        print("\n" + "="*60)
        accion = input(f"\nAcción (Número / O / .. / salir): ").strip()
        
        if accion.lower() == 'salir':
            break
        elif accion == '..':
            ruta = os.path.dirname(ruta)
        elif accion.lower() == 'o':
            ejecutar_limpieza_carpeta(ruta, archivos)
            break
        elif accion.isdigit() and int(accion) in mapeo_carpetas:
            ruta = os.path.join(ruta, mapeo_carpetas[int(accion)])
        else:
            print(f"{C_RED}Opción no válida.{C_END}")
            import time
            time.sleep(1)

def ejecutar_limpieza_carpeta(carpeta, archivos):
    # Mapeo de Categorías
    MAPEO = {
        "PDFs": [".pdf"],
        "Imagenes": [".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tiff", ".svg", ".heic"],
        "Planillas_Excel": [".csv", ".xlsx", ".xls", ".ods"],
        "Documentos": [".txt", ".docx", ".doc", ".odt", ".rtf"],
        "Comprimidos": [".zip", ".rar", ".7z", ".tar", ".gz", ".cab"]
    }
    
    # Excluir el propio script si se ejecuta desde allí
    archivos = [a for a in archivos if os.path.abspath(os.path.join(carpeta, a)) != os.path.abspath(__file__)]
    
    if not archivos:
        print(f"\n{C_YELLOW}La carpeta está limpia o no contiene archivos sueltos para organizar.{C_END}")
        presionar_enter()
        return
        
    print(f"\nSe reorganizarán {len(archivos)} archivos en la carpeta: {carpeta}")
    conf = input("¿Confirmas la acción? (S/N, por defecto S): ").strip().upper()
    if conf not in ("S", "SI", ""):
        print(f"{C_YELLOW}Operación cancelada.{C_END}")
        presionar_enter()
        return
        
    print(f"{C_YELLOW}Organizando archivos...{C_END}")
    reubicados = 0
    for file_name in archivos:
        full_path = os.path.join(carpeta, file_name)
        ext = os.path.splitext(file_name)[1].lower()
        categoria = "Otros"
        
        # Buscar categoría
        for cat, extensiones in MAPEO.items():
            if ext in extensiones:
                categoria = cat
                break
                
        # Crear subdirectorio si no existe
        subdir = os.path.join(carpeta, categoria)
        os.makedirs(subdir, exist_ok=True)
        
        # Mover archivo
        dest = os.path.join(subdir, file_name)
        try:
            shutil.move(full_path, dest)
            reubicados += 1
        except Exception as e:
            print(f"Error al mover {file_name}: {e}")
            
    print(f"\n{C_GREEN}¡Completado!{C_END} Se han clasificado y movido {reubicados} archivos con éxito en subcarpetas.")
    presionar_enter()
    presionar_enter()

# =====================================================================
# SECCIÓN 4: EXPLORADOR DE CARPETAS
# =====================================================================
def explorar_carpeta():
    print_header("EXPLORADOR INTERACTIVO DE CARPETAS")
    print("Esta herramienta te permite navegar por directorios de tu PC")
    print("para ver su estructura y qué archivos contienen sin salir de la consola.")
    
    ruta = input("\nRuta inicial (deja vacío para la carpeta actual): ").strip().strip('"')
    if not ruta:
        ruta = os.getcwd()
        
    while True:
        ruta = os.path.abspath(ruta)
        if not os.path.exists(ruta) or not os.path.isdir(ruta):
            print(f"{C_RED}La ruta '{ruta}' no es una carpeta válida.{C_END}")
            presionar_enter()
            return
            
        print_header(f"Explorando: {ruta}")
        
        try:
            items = os.listdir(ruta)
        except Exception as e:
            print(f"{C_RED}Error al acceder a la carpeta: {e}{C_END}")
            print(f"{C_YELLOW}Regresando al directorio padre...{C_END}")
            import time
            time.sleep(2)
            ruta = os.path.dirname(ruta)
            continue
            
        directorios = []
        archivos = []
        for item in items:
            # Evitar carpetas del sistema ocultas
            if item.startswith('.') or item.startswith('$'):
                continue
            full_path = os.path.join(ruta, item)
            try:
                if os.path.isdir(full_path):
                    directorios.append(item)
                else:
                    archivos.append(item)
            except Exception:
                pass
                
        directorios.sort(key=lambda x: x.lower())
        archivos.sort(key=lambda x: x.lower())
        
        idx = 1
        mapeo_items = {} # Mapea número -> (tipo, nombre)
        
        print(f" {C_YELLOW}[..]{C_END} (Subir de nivel)")
        
        if directorios:
            print(f"\n{C_CYAN}--- Carpetas ({len(directorios)}) ---{C_END}")
            for d in directorios:
                print(f"  {C_GREEN}{idx}.{C_END} [CARPETA] {d}")
                mapeo_items[idx] = ("dir", d)
                idx += 1
                
        if archivos:
            print(f"\n{C_CYAN}--- Archivos ({len(archivos)}) ---{C_END}")
            for file in archivos:
                try:
                    size = os.path.getsize(os.path.join(ruta, file)) / 1024
                    size_str = f"({size:.1f} KB)"
                except Exception:
                    size_str = ""
                print(f"  {C_GREEN}{idx}.{C_END} [ARCHIVO] {file} {C_YELLOW}{size_str}{C_END}")
                mapeo_items[idx] = ("file", file)
                idx += 1
                
        print("\n" + "="*60)
        print(f"{C_BOLD}Acciones disponibles:{C_END}")
        print(f"  - Escribe el número de una {C_GREEN}Carpeta{C_END} para entrar.")
        print(f"  - Escribe {C_RED}papelera [número]{C_END} para enviar un archivo/carpeta a la papelera.")
        print(f"  - Escribe {C_YELLOW}..{C_END} para subir de nivel.")
        print(f"  - Escribe {C_RED}salir{C_END} para volver.")
        
        accion = input(f"\nAcción: ").strip()
        
        if accion.lower() == 'salir':
            break
        elif accion == '..':
            ruta = os.path.dirname(ruta)
        elif accion.lower().startswith("papelera "):
            num_str = accion[9:].strip()
            if num_str.isdigit() and int(num_str) in mapeo_items:
                tipo, nombre = mapeo_items[int(num_str)]
                target_path = os.path.join(ruta, nombre)
                confirm = input(f"¿Seguro que deseas enviar '{nombre}' a la papelera de reciclaje? (S/N): ").strip().upper()
                if confirm in ("S", "SI"):
                    try:
                        if send2trash:
                            send2trash.send2trash(target_path)
                            print(f"{C_GREEN}¡Completado! '{nombre}' fue enviado a la Papelera de Reciclaje.{C_END}")
                        else:
                            if os.path.isdir(target_path):
                                shutil.rmtree(target_path)
                            else:
                                os.remove(target_path)
                            print(f"{C_GREEN}¡Completado! '{nombre}' fue eliminado permanentemente.{C_END}")
                    except Exception as e:
                        print(f"{C_RED}Error al eliminar: {e}{C_END}")
            else:
                print(f"{C_RED}Número de carpeta o archivo no válido.{C_END}")
            import time
            time.sleep(2)
        elif accion.isdigit() and int(accion) in mapeo_items and mapeo_items[int(accion)][0] == "dir":
            ruta = os.path.join(ruta, mapeo_items[int(accion)][1])
        else:
            print(f"{C_RED}Opción no válida.{C_END}")
            import time
            time.sleep(1.5)

# =====================================================================
# MENÚ PRINCIPAL
# =====================================================================
def main():
    # Iniciar comprobacion de actualizaciones en segundo plano de forma no bloqueante
    try:
        t = threading.Thread(target=verificar_actualizacion_segundo_plano)
        t.daemon = True
        t.start()
    except Exception:
        pass

    while True:
        print_header()
        print(f"Selecciona una categoría de herramientas:")
        print(f" {C_GREEN}1.{C_END} Gestión de Archivos PDF (Unificar, Separar, Comprimir, OCR)")
        print(f" {C_GREEN}2.{C_END} Conversor de Formatos (Imágenes, Excel, CSV, etc.)")
        print(f" {C_GREEN}3.{C_END} Organización de Oficina (Renombrado, Limpiar carpetas)")
        print(f" {C_GREEN}4.{C_END} Explorador interactivo de Carpetas")
        print(f" {C_GREEN}5.{C_END} Buscar actualizaciones del programa")
        print(f" {C_GREEN}6.{C_END} Salir")
        
        opcion = input(f"\nOpción (1-6): ").strip()
        
        if opcion == "1":
            menu_pdf()
        elif opcion == "2":
            conversor_formatos()
        elif opcion == "3":
            menu_organizacion()
        elif opcion == "4":
            explorar_carpeta()
        elif opcion == "5":
            comprobar_actualizaciones(silencioso=False)
            presionar_enter()
        elif opcion == "6":
            print(f"\n{C_CYAN}¡Gracias por usar Toolfast! Hasta luego.{C_END}")
            break
        else:
            print(f"{C_RED}Opción no válida. Ingresa un número entre 1 y 6.{C_END}")
            presionar_enter()

if __name__ == "__main__":
    main()
